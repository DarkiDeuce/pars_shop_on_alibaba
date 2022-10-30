import time
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup

file = Workbook()
file['Sheet'].title = 'Товары'
work = file['Товары']

work['A1'] = 'Имя товара'
work['B1'] = 'Ссылка на товар'
work['C1'] = 'Имя файла скриншота'

strings = 1

domen = 'https://ru1358651821imqu.trustpass.alibaba.com/'

def full_screenshot(driver, name_screenshot):

        # chrome_options = Options()
        # chrome_options.add_argument('--headless')
        # chrome_options.add_argument('--start-maximized')
        # driver = webdriver.Chrome(options=chrome_options)
        # driver.get(url=url)
        # time.sleep(2)

        #driver.execute_script("return document.scrollingElement.scrollHeight;") Получение элемента с самой длинной высотой - Справка

    ele = driver.find_element("xpath", '/html/body')
    total_height = ele.size["height"] # + 2000

    driver.set_window_size(1920, total_height)
    time.sleep(2)
    driver.save_screenshot(name_screenshot)


def list_url_products(url):
    link_products = []

    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml')

    card = soup.find_all('a', class_='product-image')

    for i in card:
        link_products.append('https:'+i.get('href'))

    return link_products

def continue_pars():
    list_product = []
    wb = load_workbook(filename='C:\\Users\\User\\Desktop\\All\\Activity\\Python\\Задачи\\ru1358651821imqu\\Таблица товаров.xlsx')
    sheet_ranges = wb['Товары']
    for product in range(2, sheet_ranges.max_row+1):
        list_product.append(sheet_ranges[f'A{product}'].value)

    return list_product

def product_information(list_url, strings, point_pars):
    for url in list_url:
        try:
            strings = strings + 1
            chromOptions = Options()
            chromOptions.add_argument('--headless')
            chromOptions.add_argument('--start-maximized')

            driver = webdriver.Chrome(options=chromOptions)

            driver.get(url)

            response = driver.page_source

            soup = BeautifulSoup(response, 'lxml')

            name_product = soup.find('div', class_='product-title').text
            name_file_screenshot = "C:\\Users\\User\\Desktop\\All\\Activity\\Python\\Задачи\\ru1358651821imqu\\products\\" + name_product + ".png"

            if name_product in point_pars:
                print('Такой товар уже есть')
                continue
            else:
                full_screenshot(driver, name_file_screenshot)

                file = load_workbook('C:\\Users\\User\\Desktop\\All\\Activity\\Python\\Задачи\\ru1358651821imqu\\Таблица товаров.xlsx')
                work = file['Товары']

                work[f'A{strings}'] = name_product
                work[f'B{strings}'] = url
                work[f'C{strings}'] = name_product + '.png'

                file.save('C:\\Users\\User\\Desktop\\All\\Activity\\Python\\Задачи\\ru1358651821imqu\\Таблица товаров.xlsx')

        finally:
            driver.close()
            driver.quit()

def main():
    list_url = list_url_products(domen)
    point_pars = continue_pars()
    product_information(list_url, strings, point_pars)

if __name__ == '__main__':
    main()

